import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.cluster import KMeans
from kneed import KneeLocator
from sklearn.metrics import silhouette_score
import sys
import os
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from PIL import Image
import tempfile

# Configuration du style matplotlib
plt.style.use('default')
plt.rcParams['figure.figsize'] = (12, 8)
plt.rcParams['axes.grid'] = True
plt.rcParams['axes.spines.top'] = False
plt.rcParams['axes.spines.right'] = False
plt.rcParams['figure.facecolor'] = 'white'
plt.rcParams['axes.facecolor'] = 'white'
sns.set_palette("husl")

def load_data(file_path):
    """Load and validate the Excel file."""
    try:
        df = pd.read_excel(file_path)
        required_columns = [
            'Customer ID', 'Age', 'Gender', 'Item Purchased', 'Category',
            'Purchase Amount (USD)', 'Location', 'Size', 'Color', 'Season',
            'Review Rating', 'Subscription Status', 'Payment Method',
            'Shipping Type', 'Discount Applied', 'Promo Code Used',
            'Previous Purchases', 'Frequency of Purchases'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
        
        return df
    except Exception as e:
        raise Exception(f"Error loading data: {str(e)}")

def preprocess_data(df):
    """Preprocess the data for clustering."""
    df_processed = df.copy()
    
    numerical_features = ['Age', 'Purchase Amount (USD)', 'Review Rating', 
                         'Previous Purchases']
    
    categorical_features = ['Gender', 'Category', 'Location', 'Size', 'Color',
                          'Season', 'Subscription Status', 'Payment Method',
                          'Shipping Type', 'Frequency of Purchases']
    
    encoders = {}
    for feature in categorical_features:
        le = LabelEncoder()
        df_processed[feature] = le.fit_transform(df_processed[feature].astype(str))
        encoders[feature] = le
    
    df_processed['Discount Applied'] = (df_processed['Discount Applied'] == 'Yes').astype(int)
    df_processed['Promo Code Used'] = (df_processed['Promo Code Used'] == 'Yes').astype(int)
    
    features_to_scale = numerical_features + categorical_features + ['Discount Applied', 'Promo Code Used']
    scaler = StandardScaler()
    df_processed[features_to_scale] = scaler.fit_transform(df_processed[features_to_scale])
    
    return df_processed, features_to_scale, encoders

def save_plot_to_temp(fig, prefix='plot'):
    """Save a matplotlib figure to a temporary file."""
    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
        fig.savefig(tmp.name)
        return tmp.name

def find_optimal_clusters(data, max_clusters=10):
    """Find the optimal number of clusters using the elbow method and silhouette score."""
    inertias = []
    silhouette_scores = []
    K = range(2, max_clusters + 1)
    
    for k in K:
        kmeans = KMeans(n_clusters=k, random_state=42)
        kmeans.fit(data)
        inertias.append(kmeans.inertia_)
        silhouette_scores.append(silhouette_score(data, kmeans.labels_))
    
    kl = KneeLocator(K, inertias, curve='convex', direction='decreasing')
    elbow_k = kl.elbow
    silhouette_k = K[np.argmax(silhouette_scores)]
    
    # Create plots
    fig = plt.figure(figsize=(10, 5))
    plt.subplot(1, 2, 1)
    plt.plot(K, inertias, 'bx-')
    plt.xlabel('Nombre de clusters (k)')
    plt.ylabel('Inertie')
    plt.title('Méthode du coude')    
    if elbow_k:
        plt.axvline(x=elbow_k, color='r', linestyle='--', label=f'Point optimal (k={elbow_k})')
        plt.legend()
    
    plt.subplot(1, 2, 2)
    plt.plot(K, silhouette_scores, 'rx-')
    plt.xlabel('Nombre de clusters (k)')
    plt.ylabel('Score de silhouette')
    plt.title('Score de silhouette par nombre de clusters')
    plt.axvline(x=silhouette_k, color='b', linestyle='--', label=f'Optimal (k={silhouette_k})')
    plt.legend()
    plt.tight_layout()
    
    # Save plot and clear figure
    plot_path = save_plot_to_temp(fig, 'elbow')
    plt.close(fig)
    
    # Prendre le maximum entre les deux méthodes pour une segmentation plus détaillée
    optimal_k = max(elbow_k, silhouette_k) if elbow_k else silhouette_k
    
    return optimal_k, plot_path

def create_distribution_plots(df, clusters):
    """Create distribution plots."""
    fig = plt.figure(figsize=(15, 5))
    
    plt.subplot(1, 3, 1)
    cluster_sizes = pd.Series(clusters).value_counts().sort_index()
    plt.pie(cluster_sizes.values, labels=[f'Cluster {i}' for i in cluster_sizes.index],
            autopct='%1.1f%%', startangle=90)
    plt.title('Répartition des clusters')
    
    plt.subplot(1, 3, 2)
    sns.boxplot(data=df, x='Cluster', y='Age')
    plt.xlabel('Cluster')
    plt.ylabel('Âge')
    plt.title('Distribution des âges par cluster')
    
    plt.subplot(1, 3, 3)
    sns.boxplot(data=df, x='Cluster', y='Purchase Amount (USD)')
    plt.xlabel('Cluster')
    plt.ylabel('Montant d\'achat (USD)')
    plt.title('Distribution des montants d\'achat par cluster')
    
    plt.tight_layout()
    
    # Save plot and clear figure
    plot_path = save_plot_to_temp(fig, 'distributions')
    plt.close(fig)
    
    return plot_path

def perform_clustering(data, n_clusters):
    """Perform KMeans clustering with the optimal number of clusters."""
    kmeans = KMeans(n_clusters=n_clusters, random_state=42)
    clusters = kmeans.fit_predict(data)
    return clusters

def analyze_clusters(df_original, df_processed, clusters, features, encoders):
    """Generate insights about each cluster using original (non-standardized) values."""
    insights = []
    
    for cluster in range(len(set(clusters))):
        cluster_data = df_original[df_original['Cluster'] == cluster]
        cluster_processed = df_processed[df_processed['Cluster'] == cluster]
        cluster_size = len(cluster_data)
        cluster_percent = (cluster_size / len(df_original)) * 100
        
        # Calculate distinctive features using standardized data
        cluster_means = cluster_processed[features].mean()
        distinctive_features = (cluster_means - df_processed[features].mean()).abs().nlargest(5)
        
        insight = f"Cluster {cluster}:\n"
        insight += f"Taille: {cluster_size} clients ({cluster_percent:.1f}%)\n"
        insight += "Caractéristiques principales:\n"
        
        # Caractéristiques numériques
        insight += f"- Âge moyen: {cluster_data['Age'].mean():.1f} ans\n"
        insight += f"- Montant d'achat moyen: {cluster_data['Purchase Amount (USD)'].mean():.2f} USD\n"
        insight += f"- Note moyenne: {cluster_data['Review Rating'].mean():.1f}/5\n"
        insight += f"- Nombre d'achats précédents: {cluster_data['Previous Purchases'].mean():.1f}\n"
        
        # Caractéristiques catégorielles les plus fréquentes
        categorical_features = ['Category', 'Season', 'Payment Method', 'Frequency of Purchases']
        for feature in categorical_features:
            most_common = cluster_data[feature].mode()[0]
            percent = (cluster_data[feature] == most_common).mean() * 100
            insight += f"- {feature} principal: {most_common} ({percent:.1f}%)\n"
        
        # Utilisation des promotions
        promo_rate = cluster_data['Promo Code Used'].value_counts(normalize=True).get('Yes', 0) * 100
        discount_rate = cluster_data['Discount Applied'].value_counts(normalize=True).get('Yes', 0) * 100
        insight += f"- Utilisation codes promo: {promo_rate:.1f}%\n"
        insight += f"- Application réductions: {discount_rate:.1f}%\n"
        
        insights.append(insight)
    
    return insights

def save_to_excel(output_path, df_original, insights, elbow_plot_path, dist_plot_path):
    """Save all data and visualizations to Excel."""
    # First save the data with pandas
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_original.to_excel(writer, sheet_name='Données avec clusters', index=False)
        insights_df = pd.DataFrame({'Analyse des clusters': insights})
        insights_df.to_excel(writer, sheet_name='Analyse des clusters', index=False)
    
    # Then load the workbook to add images
    wb = load_workbook(output_path)
    ws_viz = wb.create_sheet("Visualisations")
      # Add images
    img_elbow = XLImage(elbow_plot_path)
    img_dist = XLImage(dist_plot_path)
    
    # Get image sizes and calculate scaling
    pil_img_elbow = Image.open(elbow_plot_path)
    pil_img_dist = Image.open(dist_plot_path)
    
    # Calculate scaled dimensions for elbow plot
    width_scale = min(600 / pil_img_elbow.width, 1.0)
    height_scale = min(300 / pil_img_elbow.height, 1.0)
    scale = min(width_scale, height_scale)
    img_elbow.width = int(pil_img_elbow.width * scale)
    img_elbow.height = int(pil_img_elbow.height * scale)
    
    # Calculate scaled dimensions for distribution plot
    width_scale = min(800 / pil_img_dist.width, 1.0)
    height_scale = min(300 / pil_img_dist.height, 1.0)
    scale = min(width_scale, height_scale)
    img_dist.width = int(pil_img_dist.width * scale)
    img_dist.height = int(pil_img_dist.height * scale)
    
    # Clean up PIL Images
    pil_img_elbow.close()
    pil_img_dist.close()
    
    # Position images
    ws_viz.add_image(img_elbow, 'A1')
    ws_viz.add_image(img_dist, 'A20')
    
    # Save workbook
    wb.save(output_path)

def create_trends_plots(df):
    """Create various plots for consumption trends analysis."""
    if 'Cluster' not in df.columns:
        raise ValueError("DataFrame must contain 'Cluster' column")
    
    plots = []
    
    # 1. Category analysis
    fig, ax = plt.subplots(figsize=(12, 6))
    category_cluster = pd.crosstab(df['Category'], df['Cluster'])
    category_cluster.plot(kind='bar', stacked=True, ax=ax)
    plt.title('Répartition des catégories par cluster')
    plt.xlabel('Catégorie')
    plt.ylabel('Nombre d\'achats')
    plt.legend(title='Cluster')
    plt.xticks(rotation=45)
    plots.append(save_plot_to_temp(fig, 'categories'))
    plt.close(fig)

    # 2. Seasonal analysis
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
    season_cluster = pd.crosstab(df['Season'], df['Cluster'], normalize='columns') * 100
    season_cluster.plot(kind='bar', ax=ax1)
    ax1.set_title('Répartition saisonnière par cluster')
    ax1.set_xlabel('Saison')
    ax1.set_ylabel('Pourcentage')
    ax1.legend(title='Cluster')
    ax1.tick_params(axis='x', rotation=45)

    sns.boxplot(data=df, x='Season', y='Purchase Amount (USD)', ax=ax2)
    ax2.set_title('Montant des achats par saison')
    ax2.set_xlabel('Saison')
    ax2.set_ylabel('Montant (USD)')
    ax2.tick_params(axis='x', rotation=45)
    plt.tight_layout()
    plots.append(save_plot_to_temp(fig, 'seasonal'))
    plt.close(fig)

    # 3. Customer behavior analysis
    fig, axes = plt.subplots(2, 2, figsize=(15, 12))
    
    sns.boxplot(data=df, x='Cluster', y='Previous Purchases', ax=axes[0, 0])
    axes[0, 0].set_title('Nombre d\'achats précédents par cluster')
    
    freq_cluster = pd.crosstab(df['Frequency of Purchases'], df['Cluster'], normalize='columns') * 100
    freq_cluster.plot(kind='bar', ax=axes[0, 1])
    axes[0, 1].set_title('Fréquence d\'achat par cluster')
    axes[0, 1].tick_params(axis='x', rotation=45)
    
    sns.boxplot(data=df, x='Cluster', y='Review Rating', ax=axes[1, 0])
    axes[1, 0].set_title('Satisfaction client par cluster')
    
    subscription_cluster = pd.crosstab(df['Subscription Status'], df['Cluster'], normalize='columns') * 100
    subscription_cluster.plot(kind='bar', ax=axes[1, 1])
    axes[1, 1].set_title('Statut d\'abonnement par cluster')
    axes[1, 1].tick_params(axis='x', rotation=45)
    
    plt.tight_layout()
    plots.append(save_plot_to_temp(fig, 'behavior'))
    plt.close(fig)

    # 4. Payment and promotions analysis
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
    
    payment_cluster = pd.crosstab(df['Payment Method'], df['Cluster'], normalize='columns') * 100
    payment_cluster.plot(kind='bar', stacked=True, ax=ax1)
    ax1.set_title('Modes de paiement par cluster')
    ax1.set_xlabel('Méthode de paiement')
    ax1.set_ylabel('Pourcentage')
    ax1.legend(title='Cluster')
    ax1.tick_params(axis='x', rotation=45)
    
    # Calculate promotion usage percentages
    promo_data = pd.DataFrame()
    promo_data['Code promo'] = df.groupby('Cluster')['Promo Code Used'].apply(
        lambda x: (x == 'Yes').mean() * 100
    )
    promo_data['Réduction'] = df.groupby('Cluster')['Discount Applied'].apply(
        lambda x: (x == 'Yes').mean() * 100
    )
    
    promo_data.plot(kind='bar', ax=ax2)
    ax2.set_title('Utilisation des promotions par cluster')
    ax2.set_xlabel('Cluster')
    ax2.set_ylabel('Pourcentage d\'utilisation (%)')
    ax2.legend(title='Type de promotion')
    plt.tight_layout()
    plots.append(save_plot_to_temp(fig, 'payment_promo'))
    plt.close(fig)
    
    # 5. Correlation matrix
    fig, ax = plt.subplots(figsize=(10, 8))
    features_corr = ['Age', 'Purchase Amount (USD)', 'Review Rating', 'Previous Purchases']
    df_corr = df.copy()
    df_corr['Discount Applied'] = (df_corr['Discount Applied'] == 'Yes').astype(int)
    df_corr['Promo Code Used'] = (df_corr['Promo Code Used'] == 'Yes').astype(int)
    features_corr += ['Discount Applied', 'Promo Code Used']
    
    corr_matrix = df_corr[features_corr].corr()
    sns.heatmap(corr_matrix, annot=True, cmap='coolwarm', center=0, ax=ax)
    plt.title('Corrélations entre les variables clés')
    plt.tight_layout()
    plots.append(save_plot_to_temp(fig, 'correlations'))
    plt.close(fig)

    return plots



def analyze_customer_data(input_file, output_file):
    """Main function to analyze customer data and generate reports."""
    # Load and process data
    df_original = load_data(input_file)
    df_processed, features, encoders = preprocess_data(df_original)
    
    # Find optimal number of clusters and perform clustering
    optimal_k, elbow_plot_path = find_optimal_clusters(df_processed[features])
    clusters = perform_clustering(df_processed[features], optimal_k)
    
    # Add cluster assignments to both dataframes
    df_original['Cluster'] = clusters
    df_processed['Cluster'] = clusters
    
    # Create distribution plots using original data
    dist_plot_path = create_distribution_plots(df_original, clusters)
    
    # Generate insights using both original and processed data
    insights = analyze_clusters(df_original, df_processed, clusters, features, encoders)
    
    # Create visualizations using original data
    trends_plot_paths = create_trends_plots(df_original)
    all_plot_paths = [elbow_plot_path, dist_plot_path] + trends_plot_paths

    # Export results to Excel
    export_results_to_excel(
        output_file,
        df_original,
        insights,
        all_plot_paths
    )
    
    # Clean up temporary files
    for path in all_plot_paths:
        os.remove(path)
        
    return True  # Success

def export_results_to_excel(output_file, df_original, insights, plot_paths):
    """Export results to Excel with proper error handling."""
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Export the processed data with clusters
            df_original.to_excel(writer, sheet_name='Données analysées', index=False)
            # Export and format insights
            ws_insights = writer.book.create_sheet('Insights', 1)
            
            # Format insights as a proper DataFrame
            formatted_insights = []
            for insight in insights:
                # Split the insight into lines and process each line
                lines = insight.split('\n')
                cluster_title = lines[0].strip(':')  # "Cluster X"
                size_info = lines[1]  # "Taille: X clients (Y%)"
                
                # Process the characteristics
                characteristics = {}
                for line in lines[3:]:  # Skip "Caractéristiques principales:"
                    if line.startswith('- ') and ':' in line:
                        key, value = line[2:].split(':', 1)
                        characteristics[key.strip()] = value.strip()
                
                # Add to formatted insights
                formatted_insights.append({
                    'Cluster': cluster_title,
                    'Taille du segment': size_info,
                    **characteristics
                })
            
            # Convert to DataFrame for better presentation
            insights_df = pd.DataFrame(formatted_insights)
            
            # Write to Excel with good formatting
            insights_df.to_excel(writer, sheet_name='Insights', index=False)
            
            # Get the worksheet for formatting
            ws_insights = writer.book['Insights']
            
            # Format headers
            for cell in ws_insights[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Format data cells
            for row in ws_insights.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Adjust column widths
            for column in ws_insights.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
                ws_insights.column_dimensions[column[0].column_letter].width = adjusted_width
                
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws_insights.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Create visualizations sheet and continue with plot exports
            ws_viz = writer.book.create_sheet('Visualisations')
            
            # Define plot titles
            titles = {
                'elbow': 'Méthode du coude pour le choix du nombre de clusters',
                'silhouette': 'Score de silhouette par cluster',
                'cluster_distribution': 'Distribution des clusters',
                'cluster_profiles': 'Profils des clusters',
                'seasonal_trends': 'Tendances saisonnières par cluster',
                'category_preferences': 'Préférences de catégories par cluster',
                'correlation_heatmap': 'Matrice de corrélation des variables',
                'customer_behavior': 'Analyse des comportements clients',
                'distributions': 'Distribution des clusters et variables clés',
                'categories': 'Analyse des catégories par cluster',
                'seasonal': 'Analyse saisonnière',
                'behavior': 'Analyse comportementale',
                'payment_promo': 'Analyse des paiements et promotions',
                'correlations': 'Corrélations entre variables'
            }
            
            # Add plots to the visualization sheet
            current_row = 1
            for plot_path in plot_paths:
                if os.path.exists(plot_path):
                    # Add title
                    plot_name = os.path.splitext(os.path.basename(plot_path))[0]
                    title = titles.get(plot_name, 'Visualisation')
                    ws_viz.cell(row=current_row, column=1, value=title)
                    current_row += 1
                    
                    # Open and process image
                    try:
                        with Image.open(plot_path) as pil_img:
                            # Calculate dimensions
                            width_scale = min(600 / pil_img.width, 1.0)
                            height_scale = min(400 / pil_img.height, 1.0)
                            scale = min(width_scale, height_scale)
                            
                            img_width = int(pil_img.width * scale)
                            img_height = int(pil_img.height * scale)
                            
                            # Create Excel image
                            xl_img = XLImage(plot_path)
                            xl_img.width = img_width
                            xl_img.height = img_height
                            
                            # Add image to worksheet
                            ws_viz.add_image(xl_img, f'A{current_row}')
                            
                            # Update row position for next image
                            current_row += int(img_height / 15) + 3  # Add spacing between plots
                    except Exception as e:
                        print(f"Warning: Could not add plot {plot_path}: {str(e)}")
                        current_row += 5  # Skip some rows if image fails
            
            # Auto-adjust column widths for data sheets only
            for sheet_name in ['Données analysées', 'Insights']:
                if sheet_name in writer.book.sheetnames:
                    ws = writer.book[sheet_name]
                    for column in ws.columns:
                        max_length = 0
                        column = list(column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    except Exception as e:
        raise Exception(f"Error exporting to Excel: {str(e)}")

def analyze_customer_data(input_file, output_file):
    """Main function to analyze customer data and generate reports."""
    try:
        print("Chargement des données...")
        df_original = load_data(input_file)
        
        print("Préparation des données...")
        df_processed, features, encoders = preprocess_data(df_original)
        
        print("Recherche du nombre optimal de clusters...")
        optimal_k, elbow_plot_path = find_optimal_clusters(df_processed[features])
        print(f"Nombre optimal de clusters trouvé: {optimal_k}")
        
        print("Exécution du clustering...")
        clusters = perform_clustering(df_processed[features], optimal_k)
        
        # Add cluster assignments to both dataframes
        df_original['Cluster'] = clusters
        df_processed['Cluster'] = clusters
        
        print("Création des visualisations...")
        # Create distribution plots using original data
        dist_plot_path = create_distribution_plots(df_original, clusters)
        
        # Generate insights using both original and processed data
        insights = analyze_clusters(df_original, df_processed, clusters, features, encoders)
        
        # Create visualizations using original data
        trends_plot_paths = create_trends_plots(df_original)
        all_plot_paths = [elbow_plot_path, dist_plot_path] + trends_plot_paths

        print("Sauvegarde des résultats...")
        # Export results to Excel
        export_results_to_excel(
            output_file,
            df_original,
            insights,
            all_plot_paths
        )
        
        print("Nettoyage des fichiers temporaires...")
        # Clean up temporary files
        for path in all_plot_paths:
            try:
                if os.path.exists(path):
                    os.remove(path)
            except Exception as e:
                print(f"Warning: Could not remove temporary file {path}: {str(e)}")
                
        return True  # Success
        
    except Exception as e:
        print(f"Erreur lors de l'analyse: {str(e)}")
        return False

def main():
    try:
        if len(sys.argv) != 2:
            raise ValueError("Please provide the Excel file path as an argument")
        
        input_file = sys.argv[1]
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"File not found: {input_file}")
        
        # Ensure output directory is writable
        output_file = os.path.splitext(input_file)[0] + '_with_clusters.xlsx'
        output_dir = os.path.dirname(output_file)
        
        if not os.access(output_dir, os.W_OK):
            # Try using temp directory if original location is not writable
            import tempfile
            output_file = os.path.join(tempfile.gettempdir(), 
                                     os.path.basename(os.path.splitext(input_file)[0]) + '_with_clusters.xlsx')
            print(f"Original directory not writable, saving to: {output_file}")
        
        success = analyze_customer_data(input_file, output_file)
        
        if success:
            print(f"Analyse terminée avec succès!")
            print(f"Résultats sauvegardés dans: {output_file}")
        else:
            print("Échec de l'analyse. Vérifiez les messages d'erreur ci-dessus.")
            sys.exit(1)
        
    except Exception as e:
        print(f"Erreur fatale: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()